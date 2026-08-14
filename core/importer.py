# -*- coding: utf-8 -*-
"""
Importador de la "Consulta de información reportada por terceros" (exógena)
que el contribuyente descarga del portal de la DIAN en formato .xlsx.

Diseño:
- NUNCA se modifica el archivo original (se abre en modo lectura).
- Detecta automáticamente la fila de encabezados (no asume una posición fija),
  de modo que admite archivos con más o menos filas de metadatos/datos.
- Extrae también el encabezado informativo (año de consulta, identificación
  del consultante) para pre-llenar el perfil del contribuyente.
"""
import hashlib
from dataclasses import dataclass, field

import openpyxl

COLUMNAS_ESPERADAS = [
    "nit_reportante",
    "nombre_reportante",
    "nit_tercero",
    "nombre_tercero",
    "detalle",
    "valor",
    "uso_sugerido",
    "info_adicional",
]


@dataclass
class ExogenaImportada:
    anio_consulta: int | None = None
    tipo_documento: str | None = None
    identificacion: str | None = None
    nombre: str | None = None
    fecha_reporte: str | None = None
    hoja: str | None = None
    registros: list = field(default_factory=list)
    advertencias: list = field(default_factory=list)


def _norm(v):
    if v is None:
        return ""
    return str(v).strip()


def _es_fila_encabezado(fila) -> bool:
    textos = [_norm(c).lower() for c in fila]
    tiene_nit = any(t == "nit" for t in textos)
    tiene_valor = any("valor" in t for t in textos)
    tiene_detalle = any("detalle" in t for t in textos)
    return tiene_nit and tiene_valor and tiene_detalle


def importar_exogena(path: str) -> ExogenaImportada:
    # NOTA IMPORTANTE: se carga SIN read_only=True a propósito. El exportador
    # de la DIAN a veces escribe una etiqueta <dimension> desactualizada en el
    # XML de la hoja (declara menos filas de las que realmente tiene). El modo
    # read_only de openpyxl confía ciegamente en esa etiqueta y trunca la
    # lectura (se detectó que un archivo con 41 filas reales se leía como si
    # tuviera solo 15). Cargar el libro completo evita ese truncamiento.
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    resultado = ExogenaImportada()

    # La hoja principal suele llamarse "Reporte"; si no existe, se usa la primera.
    hoja_nombre = "Reporte" if "Reporte" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[hoja_nombre]
    resultado.hoja = hoja_nombre

    filas = list(ws.iter_rows(values_only=True))

    fila_encabezado_idx = None
    for i, fila in enumerate(filas):
        # Metadatos del encabezado informativo (año, identificación consultante)
        primera_celda = _norm(fila[0]).lower() if fila else ""
        if "año al que se refiere" in primera_celda or "ano al que se refiere" in primera_celda:
            for c in fila[1:]:
                if c not in (None, ""):
                    try:
                        resultado.anio_consulta = int(str(c).strip())
                    except ValueError:
                        pass
                    break
        elif primera_celda.startswith("tipo de documento"):
            for c in fila[1:]:
                if c not in (None, ""):
                    resultado.tipo_documento = _norm(c)
                    break
        elif primera_celda.startswith("identificación") or primera_celda.startswith("identificacion"):
            for c in fila[1:]:
                if c not in (None, ""):
                    resultado.identificacion = _norm(c)
                    break
        elif primera_celda.startswith("nombres") or "razón social" in primera_celda or "razon social" in primera_celda:
            for c in fila[1:]:
                if c not in (None, ""):
                    resultado.nombre = _norm(c)
                    break

        if _es_fila_encabezado(fila):
            fila_encabezado_idx = i
            break

    if fila_encabezado_idx is None:
        resultado.advertencias.append(
            "No se encontró la fila de encabezados esperada (NIT / Detalle / Valor). "
            "Verifique que el archivo corresponda al formato de 'Consulta de información "
            "reportada por terceros' de la DIAN."
        )
        wb.close()
        return resultado

    for i, fila in enumerate(filas[fila_encabezado_idx + 1 :], start=fila_encabezado_idx + 2):
        if all(c is None for c in fila):
            # Fin real de la tabla de la DIAN. Cualquier contenido posterior a la
            # primera fila en blanco (por ejemplo, fórmulas o notas que el usuario
            # haya agregado manualmente en el mismo archivo) se ignora a propósito:
            # esta herramienta nunca debe tratar contenido añadido por el usuario
            # en el Excel como si fuera información reportada por un tercero.
            filas_restantes = filas[i:]
            if any(any(c is not None for c in f) for f in filas_restantes):
                resultado.advertencias.append(
                    f"Se detectó contenido adicional después de la fila {i - 1} (fin de la tabla "
                    "de la DIAN) y NO se importó, para evitar confundirlo con información reportada "
                    "por terceros. Si agregó notas o fórmulas propias en este archivo, dicho contenido "
                    "fue ignorado correctamente."
                )
            break
        valor_crudo = fila[5] if len(fila) > 5 else None
        try:
            valor = float(valor_crudo) if valor_crudo not in (None, "") else None
        except (TypeError, ValueError):
            valor = None
            resultado.advertencias.append(f"Fila {i}: valor no numérico ({valor_crudo!r}), se omite.")

        detalle = _norm(fila[4]) if len(fila) > 4 else ""
        if not detalle and valor is None:
            continue

        registro = {
            "fila_excel": i,
            "nit_reportante": _norm(fila[0]) if len(fila) > 0 else "",
            "nombre_reportante": _norm(fila[1]) if len(fila) > 1 else "",
            "nit_tercero": _norm(fila[2]) if len(fila) > 2 else "",
            "nombre_tercero": _norm(fila[3]) if len(fila) > 3 else "",
            "detalle": detalle,
            "valor": valor,
            "uso_sugerido": _norm(fila[6]) if len(fila) > 6 else "",
            "info_adicional": _norm(fila[7]) if len(fila) > 7 else "",
        }
        # Hash para detectar duplicados exactos (mismo tercero+concepto+valor)
        base_hash = f"{registro['nit_reportante']}|{registro['nombre_reportante']}|{registro['detalle']}|{registro['valor']}"
        registro["hash_fila"] = hashlib.sha256(base_hash.encode("utf-8")).hexdigest()[:16]
        resultado.registros.append(registro)

    wb.close()

    if not resultado.registros:
        resultado.advertencias.append(
            "El archivo no contiene filas de datos reportadas por terceros. "
            "Esto puede significar que ningún tercero reportó información para este año, "
            "o que la consulta se generó antes de que los reportantes cargaran su información "
            "(los plazos de reporte de exógena vencen a inicios del año siguiente)."
        )

    return resultado


def detectar_duplicados(registros: list[dict]) -> list[dict]:
    vistos = {}
    for r in registros:
        h = r["hash_fila"]
        vistos.setdefault(h, []).append(r)
    duplicados = []
    for h, grupo in vistos.items():
        if len(grupo) > 1:
            duplicados.extend(grupo)
    return duplicados
